"""
export.py

Responsável por:
1. Exportar dados para Excel
2. Criar aba Dashboard
3. Inserir imagens dos gráficos
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import pandas as pd


def exportar_excel(df, df_categoria, arquivo_excel):
    """
    Exporta os dados e o resumo para um arquivo Excel.
    """

    with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Transacoes", index=False)
        df_categoria.to_excel(writer, sheet_name="Resumo por Categoria")


def adicionar_dashboard_excel(arquivo_excel, img_categoria, img_dia):
    """
    Cria aba 'Dashboard' e adiciona os gráficos salvos como imagem.
    """

    wb = load_workbook(arquivo_excel)
    ws_dashboard = wb.create_sheet("Dashboard")

    ws_dashboard.add_image(ExcelImage(img_categoria), "A1")
    ws_dashboard.add_image(ExcelImage(img_dia), "A50")

    wb.save(arquivo_excel)
